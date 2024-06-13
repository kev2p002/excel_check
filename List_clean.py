import numpy as np

# Sample list containing various values, including NaN and blank spaces
data = [1, 2, np.nan, '', 'Hello', ' ', None, 'World', 3.5, float('nan')]

# Remove NaN and blank space values
cleaned_data = [x for x in data if not (x is None or (isinstance(x, float) and np.isnan(x)) or str(x).strip() == '')]

print(cleaned_data)

color dict
import pandas as pd
from openpyxl import load_workbook

# Load the Excel file
excel_path = 'your_excel_file.xlsx'
wb = load_workbook(excel_path, data_only=True)

# Load the second sheet
sheet_name = wb.sheetnames[1]
ws = wb[sheet_name]

# Read the second sheet into a DataFrame
df = pd.read_excel(excel_path, sheet_name=sheet_name)

# Extract the color information and convert to hex values
def rgb_to_hex(rgb):
    return f'#{int(rgb[0]):02X}{int(rgb[1]):02X}{int(rgb[2]):02X}'

def get_cell_color(cell):
    color = cell.fill.start_color
    if color.type == 'rgb':
        return rgb_to_hex((int(color.rgb[2:4], 16), int(color.rgb[4:6], 16), int(color.rgb[6:8], 16)))
    else:
        return None

colors = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=3):
    for cell in row:
        colors.append(get_cell_color(cell))

# Add the color information to the DataFrame
df['colour'] = colors

# Create a dictionary of meaning:color hex code
meaning_color_dict = df.set_index('meaning')['colour'].to_dict()

# Display the DataFrame and the dictionary
print(df)
print(meaning_color_dict)

color dataframe
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Color

# Load the Excel file
excel_path = 'your_excel_file.xlsx'
wb = load_workbook(excel_path, data_only=True)

# Load the second sheet
sheet_name = wb.sheetnames[1]
ws = wb[sheet_name]

# Read the second sheet into a DataFrame
df = pd.read_excel(excel_path, sheet_name=sheet_name)

# Extract the color information and convert to hex values
def rgb_to_hex(rgb):
    return f'#{int(rgb[0]):02X}{int(rgb[1]):02X}{int(rgb[2]):02X}'

def get_cell_color(cell):
    color = cell.fill.start_color
    if color.type == 'rgb':
        return rgb_to_hex((int(color.rgb[2:4], 16), int(color.rgb[4:6], 16), int(color.rgb[6:8], 16)))
    else:
        return None

colors = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=3):
    for cell in row:
        colors.append(get_cell_color(cell))

# Add the color information to the DataFrame
df['colour'] = colors

# Display the DataFrame
print(df)
