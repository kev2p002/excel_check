from flask import Flask, render_template, request, send_file
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill

app = Flask(__name__)

@app.route('/')
def index():
    data = {
        "hello": {
            "High": ["Item1", "Item2"],
            "Low": ["Item3", "Item4", "Item5"],
            "Normal": ["Item6", "Item7", "Item8"]
        },
        "man": {
            "High": ["Item1", "Item2"],
            "Low": ["Item3"],
            "Normal": ["Item4", "Item5"]
        }
    }
    
    dataframes = {}
    pie_data = {}

    for key, value in data.items():
        rows = []
        for impact, items in value.items():
            for item in items:
                rows.append({"Impact": impact, "Item": item, "ColorClass": impact.lower()})
        df = pd.DataFrame(rows)
        dataframes[key] = df
        pie_data[key] = {
            "High": len(value["High"]),
            "Low": len(value["Low"]),
            "Normal": len(value["Normal"])
        }

    return render_template('index.html', dataframes=dataframes, pie_data=pie_data)

@app.route('/download_excel', methods=['POST'])
def download_excel():
    data = request.json
    key = data['key']
    rows = data['rows']
    df = pd.DataFrame(rows)

    output = io.BytesIO()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = key

    # Adding headers
    headers = ['Impact', 'Item', 'Interaction', 'Alerts']
    sheet.append(headers)

    # Adding data with 'Alerts' column
    for _, row in df.iterrows():
        alert_color = 'FFFFFF'  # Default white
        if row['Impact'] == 'High':
            alert_color = 'FF0000'  # Red
        elif row['Impact'] == 'Low':
            alert_color = 'FFFF00'  # Yellow
        elif row['Impact'] == 'Normal':
            alert_color = 'FFA500'  # Orange
        sheet.append([row['Impact'], row['Item'], row['Interaction'], ''])
        cell = sheet.cell(row=sheet.max_row, column=4)
        cell.fill = PatternFill(start_color=alert_color, end_color=alert_color, fill_type="solid")

    workbook.save(output)
    output.seek(0)

    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=f'{key}.xlsx')

if __name__ == '__main__':
    app.run(debug=True)

<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Expandable Cards with DataTables</title>
    <link href="https://cdn.datatables.net/1.11.5/css/jquery.dataTables.min.css" rel="stylesheet">
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.1.3/dist/css/bootstrap.min.css" rel="stylesheet">
    <script src="https://code.jquery.com/jquery-3.6.0.min.js"></script>
    <script src="https://cdn.datatables.net/1.11.5/js/jquery.dataTables.min.js"></script>
    <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.1.3/dist/js/bootstrap.bundle.min.js"></script>
    <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
    <style>
        .high { background-color: red; }
        .normal { background-color: orange; }
        .low { background-color: yellow; }
        .chart-container {
            position: relative;
            width: 100%;
            height: 400px; /* default height */
        }
        .dataTables_wrapper .dataTables_filter {
            margin-bottom: 15px; /* space between search bar and table */
        }
        .dataTables_wrapper .dataTables_paginate {
            margin-top: 15px; /* space between table and pagination */
        }
        .strikeout {
            text-decoration: line-through;
        }
    </style>
</head>
<body>
    <div class="container mt-5">
        {% for key, df in dataframes.items() %}
        <div class="card mb-3">
            <div class="card-header">
                <h5 class="mb-0">
                    <button class="btn btn-link" data-bs-toggle="collapse" data-bs-target="#collapse{{ loop.index }}" aria-expanded="true" aria-controls="collapse{{ loop.index }}">
                        {{ key }}
                    </button>
                </h5>
            </div>
            <div id="collapse{{ loop.index }}" class="collapse" data-bs-parent="#accordion">
                <div class="card-body">
                    <div class="row">
                        <div class="col-md-8">
                            <table class="table table-striped dataTable" id="dataTable{{ loop.index }}">
                                <thead>
                                    <tr>
                                        <th>Impact</th>
                                        <th>Item</th>
                                        <th>Interaction</th>
                                    </tr>
                                </thead>
                                <tbody>
                                    {% for row in df.itertuples() %}
                                    <tr class="{{ row.ColorClass }}">
                                        <td>{{ row.Impact }}</td>
                                        <td>{{ row.Item }}</td>
                                        <td>
                                            <select class="form-select interaction-dropdown" data-row="{{ loop.index }}" data-item="{{ row.Item }}">
                                                <option value="Select">Select</option>
                                                <option value="Major">Major</option>
                                                <option value="Minor">Minor</option>
                                                <option value="No Interaction">No Interaction</option>
                                            </select>
                                        </td>
                                    </tr>
                                    {% endfor %}
                                </tbody>
                            </table>
                            <button class="btn btn-primary mt-3 download-btn" data-key="{{ key }}" data-index="{{ loop.index }}">Download Excel</button>
                        </div>
                        <div class="col-md-4">
                            <div class="chart-container">
                                <canvas id="pieChart{{ loop.index }}"></canvas>
                            </div>
                        </div>
                    </div>
                </div>
            </div>
        </div>
        {% endfor %}
    </div>

    <script>
        $(document).ready(function() {
            var dataTables = {};
            {% for key, counts in pie_data.items() %}
            var ctx = document.getElementById('pieChart{{ loop.index }}').getContext('2d');
            var pieChart = new Chart(ctx, {
                type: 'pie',
                data: {
                    labels: ['High', 'Low', 'Normal'],
                    datasets: [{
                        data: [{{ counts['High'] }}, {{ counts['Low'] }}, {{ counts['Normal'] }}],
                        backgroundColor: ['red', 'yellow', 'orange']
                    }]
                },
                options: {
                    responsive: true,
                    maintainAspectRatio: false
                }
            });

            var dataTable = $('#dataTable{{ loop.index }}').DataTable({
                "pageLength": 5,
                "drawCallback": function(settings) {
                    adjustChartContainerHeight('{{ loop.index }}');
                }
            });
            dataTables['dataTable{{ loop.index }}'] = dataTable;

            // Adjust chart container height based on DataTable
            function adjustChartContainerHeight(index) {
                var tableHeight = $('#dataTable' + index + '_wrapper .dataTables_scrollBody').height();
                $('#collapse' + index + ' .chart-container').css('height', tableHeight);
            }

            adjustChartContainerHeight('{{ loop.index }}');

            {% endfor %}

            // Add event listener for interaction dropdown
            $(document).on('change', '.interaction-dropdown', function() {
                var row = $(this).closest('tr');
                if ($(this).val() === 'No Interaction') {
                    row.addClass('strikeout');
                } else {
                    row.removeClass('strikeout');
                }
            });

            // Add event listener for download buttons
            $(document).on('click', '.download-btn', function() {
                var key = $(this).data('key');
                var index = $(this).data('index');
                var dataTable = dataTables['dataTable' + index];

                var rows = [];
                dataTable.rows().every(function() {
                    var row = $(this.node());
                    var impact = row.find('td').eq(0).text();
                    var item = row.find('td').eq(1).text();
                    var interaction = row.find('select').val();
                    rows.push({ 'Impact': impact, 'Item': item, 'Interaction': interaction });
                });

                // Send data to Flask endpoint to generate Excel file
                $.ajax({
                    url: '/download_excel',
                    method: 'POST',
                    contentType: 'application/json',
                    data: JSON.stringify({ key: key, rows: rows }),
                    xhrFields: {
                        responseType: 'blob'
                    },
                    success: function(blob) {
                        var url = window.URL.createObjectURL(blob);
                        var a = document.createElement('a');
                        a.style.display = 'none';
                        a.href = url;
                        a.download = key + '.xlsx';
                        document.body.appendChild(a);
                        a.click();
                        window.URL.revokeObjectURL(url);
                    },
                    error: function() {
                        alert('Error downloading file.');
                    }
                });
            });
        });
    </script>
</body>
</html>
