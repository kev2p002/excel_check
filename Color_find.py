from openpyxl import load_workbook
from openpyxl.styles import Color

def rgb_to_hex(rgb):
    return f"#{rgb[2]:02x}{rgb[1]:02x}{rgb[0]:02x}".upper()

# Load the Excel workbook
wb = load_workbook('path_to_your_excel_file.xlsx')

# Select the active sheet (or specify the sheet name)
ws = wb.active

# Get the cell with the color (first column of the first row)
cell = ws.cell(row=1, column=1)

# Get the fill color of the cell
fill = cell.fill

# Check if the fill has a start_color (for solid fills)
if fill.start_color.index != '00000000':
    color = fill.start_color.rgb  # Returns 'RRGGBBAA'
    # Extract the RGB part and convert to hex
    hex_color = f"#{color[2:4]}{color[4:6]}{color[6:8]}"
    print(f"Hex color: {hex_color}")
else:
    print("The cell does not have a fill color.")

# For pattern fills (e.g., gradient), the approach might vary
