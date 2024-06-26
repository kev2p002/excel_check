from flask import Flask, render_template, request, send_file, jsonify
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment,Border, Side

app = Flask(__name__)

@app.route('/')
def index():
    data = {
        "hello": {
            "Highman": ["Item1", "Item2"],
            "Low": ["Item3", "Item4", "Item5"],
            "Normal": ["Item6", "Item7", "Item8"]
        },
        "man": {
            "Highman": ["Item1", "Item2"],
            "Low": ["Item3"],
            "Normal": ["Item4", "Item5"]
        }
    }

    impact_colors = {
        "Highman": "#FF0000",
        "Low": "#FFFF00",
        "Normal": "#FFA500"
    }

    dataframes = {}
    pie_data = {}

    for key, value in data.items():
        rows = []
        for impact, items in value.items():
            for item in items:
                rows.append({"Impact": impact, "Item": item, "ColorClass": impact.lower()})
        df = pd.DataFrame(rows)
        dataframes[key] = df
        pie_data[key] = {
            "Highman": len(value["Highman"]),
            "Low": len(value["Low"]),
            "Normal": len(value["Normal"])
        }

    return render_template('summary.html', dataframes=dataframes, pie_data=pie_data, impact_colors=impact_colors)


@app.route('/download_excel', methods=['POST'])
def download_excel():
    data = request.json
    sheets_data = data['sheets']

    # Prepare data for the summary sheet
    summary_data = {
        "High": {"count": 0, "features": []},
        "Low": {"count": 0, "features": []},
        "Normal": {"count": 0, "features": []}
    }

    for sheet_data in sheets_data:
        key = sheet_data['key']
        rows = sheet_data['rows']
        for row in rows:
            if row['Impact'] in summary_data:
                summary_data[row['Impact']]['count'] += 1
                if key not in summary_data[row['Impact']]['features']:
                    summary_data[row['Impact']]['features'].append(key)

    output = io.BytesIO()
    workbook = Workbook()
    workbook.remove(workbook.active)  # Remove default sheet

    # Create summary sheet
    summary_sheet = workbook.create_sheet(title="Summary", index=0)

    # Set background color
    fill = PatternFill(start_color="95B3D7", end_color="ADD8E6", fill_type="solid")
    
    # Set border style
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))

    # Title cell
    title_cell = summary_sheet.cell(row=1, column=1, value="IMPACT SUMMARY")
    summary_sheet.merge_cells("A1:C1")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.font = Font(size=14, bold=True)
    title_cell.fill = fill
    # title_cell.border = thin_border
    summary_sheet.column_dimensions['A'].width = 40  # Adjust the width for the first column

    # Description cell
    description_cell = summary_sheet.cell(row=2, column=1, value="Find all info related to summary of sheets here")
    summary_sheet.merge_cells("A2:C2")
    description_cell.alignment = Alignment(horizontal="center", vertical="center")
    description_cell.fill = fill
    # description_cell.border = thin_border

    # Headers row
    headers = ["Impact", "Count", "Features"]
    for col_num, header in enumerate(headers, 1):
        cell = summary_sheet.cell(row=3, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.border = thin_border
        # Set width for each header column
        if col_num == 1:
            summary_sheet.column_dimensions['A'].width = 15
        elif col_num == 2:
            summary_sheet.column_dimensions['B'].width = 10
        elif col_num == 3:
            summary_sheet.column_dimensions['C'].width = 40

    # Data rows
    for row_num, (impact, data) in enumerate(summary_data.items(), start=4):
        features_str = ", ".join(data["features"])
        row = [impact, data["count"], features_str]
        for col_num, value in enumerate(row, 1):
            cell = summary_sheet.cell(row=row_num, column=col_num, value=value)
            cell.fill = fill
            cell.border = thin_border

    # Create individual sheets
    for sheet_data in sheets_data:
        key = sheet_data['key']
        rows = sheet_data['rows']
        df = pd.DataFrame(rows)
        sheet = workbook.create_sheet(title=key)

        # Adding headers
        headers = ['Impact', 'Item', 'Interaction', 'Alerts']
        sheet.append(headers)

        # Adding data with 'Alerts' column
        for _, row in df.iterrows():
            alert_color = 'FFFFFF'  # Default white
            if row['Impact'] == 'High':
                alert_color = 'FF0000'  # Red
            elif row['Impact'] == 'Low':
                alert_color = 'FFFF00'  # Yellow
            elif row['Impact'] == 'Normal':
                alert_color = 'FFA500'  # Orange
            
            # Writing data to the sheet
            sheet.append([row['Impact'], row['Item'], row['Interaction'], ''])
            # Formatting font color based on impact
            cell = sheet.cell(row=sheet.max_row, column=1)
            cell.font = Font(color=alert_color)

    workbook.save(output)
    output.seek(0)

    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name='all_data.xlsx')

if __name__ == '__main__':
    app.run(debug=True)


<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Expandable Cards with DataTables</title>
    <link href="https://cdn.datatables.net/1.11.5/css/jquery.dataTables.min.css" rel="stylesheet">
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.1.3/dist/css/bootstrap.min.css" rel="stylesheet">
    <script src="https://code.jquery.com/jquery-3.6.0.min.js"></script>
    <script src="https://cdn.datatables.net/1.11.5/js/jquery.dataTables.min.js"></script>
    <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.1.3/dist/js/bootstrap.bundle.min.js"></script>
    <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
    <style>
        .chart-container {
            position: relative;
            width: 100%;
            height: 400px; /* default height */
        }
        .dataTables_wrapper .dataTables_filter {
            margin-bottom: 15px; /* space between search bar and table */
        }
        .dataTables_wrapper .dataTables_paginate {
            margin-top: 15px; /* space between table and pagination */
        }
        .strikeout {
            text-decoration: line-through;
        }
    </style>
</head>
<body>
    <div class="container mt-5">
        {% for key, df in dataframes.items() %}
        <div class="card mb-3">
            <div class="card-header">
                <h5 class="mb-0">
                    <button class="btn btn-link" data-bs-toggle="collapse" data-bs-target="#collapse{{ loop.index }}" aria-expanded="true" aria-controls="collapse{{ loop.index }}">
                        {{ key }}
                    </button>
                </h5>
            </div>
            <div id="collapse{{ loop.index }}" class="collapse" data-bs-parent="#accordion">
                <div class="card-body">
                    <div class="row">
                        <div class="col-md-8">
                            <table class="table table-striped dataTable" id="dataTable{{ loop.index }}">
                                <thead>
                                    <tr>
                                        <th>Impact</th>
                                        <th>Item</th>
                                        <th>Interaction</th>
                                    </tr>
                                </thead>
                                <tbody>
                                    {% for row in df.itertuples() %}
                                    <tr>
                                        <td>{{ row.Impact }}</td>
                                        <td>{{ row.Item }}</td>
                                        <td>
                                            <select class="form-select interaction-dropdown" data-row="{{ loop.index }}" data-item="{{ row.Item }}">
                                                <option value="Select">Select</option>
                                                <option value="Major">Major</option>
                                                <option value="Minor">Minor</option>
                                                <option value="No Interaction">No Interaction</option>
                                            </select>
                                        </td>
                                    </tr>
                                    {% endfor %}
                                </tbody>
                            </table>
                        </div>
                        <div class="col-md-4">
                            <div class="chart-container">
                                <canvas id="pieChart{{ loop.index }}"></canvas>
                            </div>
                        </div>
                    </div>
                </div>
            </div>
        </div>
        {% endfor %}
        <button class="btn btn-primary mt-3" id="downloadAllBtn">Download All Data</button>
    </div>

    <script>
        $(document).ready(function() {
            var dataTables = {};
            var impactColors = {{ impact_colors|tojson|safe }};
            var pieData = {{ pie_data|tojson|safe }};

            {% for key in dataframes.keys() %}
            var ctx = document.getElementById('pieChart{{ loop.index }}').getContext('2d');
            var labels = Object.keys(pieData['{{ key }}']);
            var data = Object.values(pieData['{{ key }}']);
            var backgroundColors = labels.map(function(label) {
                return impactColors[label];
            });

            var pieChart = new Chart(ctx, {
                type: 'pie',
                data: {
                    labels: labels,
                    datasets: [{
                        data: data,
                        backgroundColor: backgroundColors
                    }]
                },
                options: {
                    responsive: true,
                    maintainAspectRatio: false
                }
            });

            var dataTable = $('#dataTable{{ loop.index }}').DataTable({
                "pageLength": 5,
                "drawCallback": function(settings) {
                    adjustChartContainerHeight('{{ loop.index }}');
                    updateCellColors('{{ loop.index }}'); // Update cell colors on every draw
                }
            });
            dataTables['dataTable{{ loop.index }}'] = dataTable;

            // Adjust chart container height based on DataTable
            function adjustChartContainerHeight(index) {
                var tableHeight = $('#dataTable' + index + '_wrapper .dataTables_scrollBody').height();
                $('#collapse' + index + ' .chart-container').css('height', tableHeight);
            }

            adjustChartContainerHeight('{{ loop.index }}');

            {% endfor %}

            // Add event listener for interaction dropdown
            $(document).on('change', '.interaction-dropdown', function() {
                var row = $(this).closest('tr');
                if ($(this).val() === 'No Interaction') {
                    row.addClass('strikeout');
                } else {
                    row.removeClass('strikeout');
                }
            });

            // Add event listener for the download all button
            $('#downloadAllBtn').on('click', function() {
                var sheets = [];
                {% for key in dataframes.keys() %}
                var dataTable = dataTables['dataTable{{ loop.index }}'];
                var rows = [];
                dataTable.rows().every(function() {
                    var row = $(this.node());
                    var impact = row.find('td').eq(0).text();
                    var item = row.find('td').eq(1).text();
                    var interaction = row.find('select').val();
                    rows.push({ 'Impact': impact, 'Item': item, 'Interaction': interaction });
                });
                sheets.push({ key: '{{ key }}', rows: rows });
                {% endfor %}

                // Send data to Flask endpoint to generate Excel file
                $.ajax({
                    url: '/download_excel',
                    method: 'POST',
                    contentType: 'application/json',
                    data: JSON.stringify({ sheets: sheets }),
                    xhrFields: {
                        responseType: 'blob'
                    },
                    success: function(blob) {
                        var url = window.URL.createObjectURL(blob);
                        var a = document.createElement('a');
                        a.style.display = 'none';
                        a.href = url;
                        a.download = 'all_data.xlsx';
                        document.body.appendChild(a);
                        a.click();
                        window.URL.revokeObjectURL(url);
                    },
                    error: function() {
                        alert('Error downloading file.');
                    }
                });
            });

            // Function to update cell colors
            function updateCellColors(index) {
                $('#dataTable' + index + ' tbody tr').each(function() {
                    var impactCell = $(this).find('td').eq(0);
                    var impactText = impactCell.text();
                    var color = impactColors[impactText];
                    console.log('Impact:', impactText, 'Color:', color); // Log the impact and color to the console
                    if (color) {
                        impactCell.css('color', color);
                    }
                });
            }
        });
    </script>
</body>
</html>
