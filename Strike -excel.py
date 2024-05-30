from flask import Flask, render_template, request, send_file
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

app = Flask(__name__)

@app.route('/')
def index():
    data = {
        "hello": {
            "High": ["Item1", "Item2"],
            "Low": ["Item3", "Item4", "Item5"],
            "Normal": ["Item6", "Item7", "Item8"]
        },
        "man": {
            "High": ["Item1", "Item2"],
            "Low": ["Item3"],
            "Normal": ["Item4", "Item5"]
        }
    }

    dataframes = {}
    pie_data = {}

    for key, value in data.items():
        rows = []
        for impact, items in value.items():
            for item in items:
                rows.append({"Impact": impact, "Item": item, "ColorClass": impact.lower()})
        df = pd.DataFrame(rows)
        dataframes[key] = df
        pie_data[key] = {
            "High": len(value["High"]),
            "Low": len(value["Low"]),
            "Normal": len(value["Normal"])
        }

    return render_template('index.html', dataframes=dataframes, pie_data=pie_data)

@app.route('/download_excel', methods=['POST'])
def download_excel():
    data = request.json
    sheets_data = data['sheets']

    output = io.BytesIO()
    workbook = Workbook()
    workbook.remove(workbook.active)  # Remove default sheet

    for sheet_data in sheets_data:
        key = sheet_data['key']
        rows = sheet_data['rows']
        df = pd.DataFrame(rows)
        sheet = workbook.create_sheet(title=key)

        # Adding headers
        headers = ['Impact', 'Item', 'Interaction', 'Alerts']
        sheet.append(headers)

        # Adding data with 'Alerts' column
        for _, row in df.iterrows():
            alert_color = 'FFFFFF'  # Default white
            fill_type = "solid"  # Default fill type
            if row['Impact'] == 'High':
                alert_color = 'FF0000'  # Red
                fill_type = "darkGrid"  # Example: change fill type to dark grid for High impact
            elif row['Impact'] == 'Low':
                alert_color = 'FFFF00'  # Yellow
                fill_type = "lightHorizontal"  # Example: change fill type to light horizontal for Low impact
            elif row['Impact'] == 'Normal':
                alert_color = 'FFA500'  # Orange
                fill_type = "darkVertical"  # Example: change fill type to dark vertical for Normal impact
            
            # Append row data
            sheet.append([row['Impact'], row['Item'], row['Interaction'], ''])
            for col in range(1, 5):  # Apply fill to all columns in the row
                cell = sheet.cell(row=sheet.max_row, column=col)
                cell.fill = PatternFill(start_color=alert_color, end_color=alert_color, fill_type=fill_type)

                # Apply strikethrough if interaction is "No Interaction"
                if row['Interaction'] == 'No Interaction':
                    cell.font = Font(strike=True)

    workbook.save(output)
    output.seek(0)

    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name='all_data.xlsx')

if __name__ == '__main__':
    app.run(debug=True)
